"""Module 4 — real, filter-respecting exports (CSV / Excel / PDF) built
straight from the rows a report screen is already showing. No fixture
data: callers pass in the same `columns` + `rows` used to render the HTML
table."""
import csv
import io
from datetime import datetime

from flask import Response


def _filename(base, ext):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{stamp}.{ext}"


def export_csv(title, columns, rows, base_name="report"):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    output = buf.getvalue()
    resp = Response(output, mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename={_filename(base_name, 'csv')}"
    return resp


_INVALID_SHEET_TITLE_CHARS = r"[]:*?/\\"


def _safe_sheet_title(title):
    """openpyxl forbids \\ / * ? : [ ] in sheet titles and caps them at 31 chars."""
    cleaned = "".join(c for c in (title or "Report") if c not in _INVALID_SHEET_TITLE_CHARS)
    cleaned = cleaned.strip() or "Report"
    return cleaned[:31]


def export_excel(title, columns, rows, base_name="report"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(title)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    title_cell = ws.cell(row=1, column=1, value=title or "SATS Report")
    title_cell.font = Font(size=14, bold=True, color="0D5C3F")
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(list(columns))
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D5C3F", end_color="0D5C3F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(list(row))

    for i, col in enumerate(columns, start=1):
        max_len = max([len(str(col))] + [len(str(r[i - 1])) for r in rows]) if rows else len(str(col))
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = min(max(max_len + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f"attachment; filename={_filename(base_name, 'xlsx')}"
    return resp


def export_pdf(title, columns, rows, base_name="report", subtitle=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title or "SATS Report", styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 8))

    data = [list(columns)] + [[("" if c is None else str(c)) for c in row] for row in rows]
    if len(data) == 1:
        data.append(["No data available yet."] + [""] * (len(columns) - 1))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D5C3F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F7F5")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    resp = Response(buf.read(), mimetype="application/pdf")
    resp.headers["Content-Disposition"] = f"attachment; filename={_filename(base_name, 'pdf')}"
    return resp


def do_export(fmt, title, columns, rows, base_name="report", subtitle=None):
    fmt = (fmt or "csv").lower()
    if fmt == "excel" or fmt == "xlsx":
        return export_excel(title, columns, rows, base_name)
    if fmt == "pdf":
        return export_pdf(title, columns, rows, base_name, subtitle=subtitle)
    return export_csv(title, columns, rows, base_name)
